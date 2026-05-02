from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, Depends, Header, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from datetime import datetime
import models, schemas, auth, csv, io
from database import engine, SessionLocal
from dependencies import get_current_user, admin_only
from config import ADMIN_SECRET_CODE

models.Base.metadata.create_all(bind=engine)

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



app.mount("/static", StaticFiles(directory="frontend"), name="static")




def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/")
def root():
    return {"message": "TaskPro API is running"}

# ---------- AUTH ----------
@app.post("/signup")
def signup(user: schemas.UserCreate, db: Session = Depends(get_db)):
    existing = db.query(models.User).filter(models.User.username == user.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    if user.role == "admin":
        if not user.admin_code or user.admin_code != ADMIN_SECRET_CODE:
            raise HTTPException(status_code=403, detail="Invalid admin secret code")
    if user.role not in ("admin", "member"):
        raise HTTPException(status_code=400, detail="Invalid role")
    hashed = auth.hash_password(user.password)
    new_user = models.User(username=user.username, password=hashed, role=user.role)
    db.add(new_user)
    db.commit()
    return {"msg": "User created"}

@app.post("/login")
def login(user: schemas.LoginSchema, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.username == user.username).first()
    if not db_user or not auth.verify_password(user.password, db_user.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = auth.create_token({"user_id": db_user.id, "role": db_user.role, "username": db_user.username})
    return {"token": token, "role": db_user.role, "user_id": db_user.id, "username": db_user.username}

# ---------- USERS ----------
@app.get("/users")
def get_users(db: Session = Depends(get_db), token: str = Header(...)):
    get_current_user(token)
    users = db.query(models.User).all()
    return [{"id": u.id, "username": u.username, "role": u.role} for u in users]

@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    admin_only(current)
    if current["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.query(models.Task).filter(models.Task.assigned_to == user_id).update({"assigned_to": None})
    db.delete(user)
    db.commit()
    return {"msg": f"User '{user.username}' removed successfully"}

# ---------- USER STATS ----------
@app.get("/users/{user_id}/stats")
def get_user_stats(user_id: int, db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    if current["role"] != "admin" and current["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    tasks = db.query(models.Task).filter(models.Task.assigned_to == user_id).all()
    now = datetime.utcnow()
    result = []
    for t in tasks:
        project = db.query(models.Project).filter(models.Project.id == t.project_id).first()
        result.append({
            "id": t.id, "title": t.title, "status": t.status,
            "project_name": project.name if project else "Unknown",
            "project_id": t.project_id,
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "overdue": bool(t.due_date and t.due_date < now and t.status != "completed")
        })
    return {
        "total": len(tasks),
        "completed": len([t for t in tasks if t.status == "completed"]),
        "pending": len([t for t in tasks if t.status != "completed"]),
        "tasks": result
    }

# ---------- PROJECTS ----------
@app.post("/projects")
def create_project(project: schemas.ProjectCreate, db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create projects")
    new_project = models.Project(name=project.name, owner_id=user["user_id"])
    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    return {"id": new_project.id, "name": new_project.name, "owner_id": new_project.owner_id}

@app.get("/projects")
def get_projects(db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    if current["role"] == "admin":
        projects = db.query(models.Project).all()
    else:
        assigned_project_ids = db.query(models.Task.project_id).filter(
            models.Task.assigned_to == current["user_id"]
        ).distinct().all()
        pids = [r[0] for r in assigned_project_ids]
        projects = db.query(models.Project).filter(models.Project.id.in_(pids)).all()
    return [{"id": p.id, "name": p.name, "owner_id": p.owner_id} for p in projects]

@app.get("/projects/{project_id}/detail")
def get_project_detail(project_id: int, db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current["role"] == "admin":
        tasks = db.query(models.Task).filter(models.Task.project_id == project_id).all()
    else:
        tasks = db.query(models.Task).filter(
            models.Task.project_id == project_id,
            models.Task.assigned_to == current["user_id"]
        ).all()
    now = datetime.utcnow()
    user_ids = list(set(t.assigned_to for t in tasks if t.assigned_to))
    members = []
    for uid in user_ids:
        u = db.query(models.User).filter(models.User.id == uid).first()
        if u:
            user_tasks = [t for t in tasks if t.assigned_to == uid]
            members.append({
                "id": u.id, "username": u.username, "role": u.role,
                "task_count": len(user_tasks),
                "done_count": len([t for t in user_tasks if t.status == "completed"])
            })
    task_list = []
    for t in tasks:
        u = db.query(models.User).filter(models.User.id == t.assigned_to).first()
        task_list.append({
            "id": t.id, "title": t.title, "status": t.status,
            "assigned_to": t.assigned_to,
            "assigned_username": u.username if u else "Unassigned",
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "overdue": bool(t.due_date and t.due_date < now and t.status != "completed")
        })
    total = len(tasks)
    completed = len([t for t in tasks if t.status == "completed"])
    return {
        "id": project.id, "name": project.name, "owner_id": project.owner_id,
        "total_tasks": total, "completed_tasks": completed,
        "progress_pct": round((completed / total * 100) if total > 0 else 0),
        "members": members, "tasks": task_list
    }

# ---------- TASKS ----------
@app.post("/tasks")
def create_task(task: schemas.TaskCreate, db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create tasks")
    project = db.query(models.Project).filter(models.Project.id == task.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    new_task = models.Task(title=task.title, due_date=task.due_date,
                           assigned_to=task.assigned_to, project_id=task.project_id, status="pending")
    db.add(new_task)
    db.commit()
    db.refresh(new_task)
    assigned_user = db.query(models.User).filter(models.User.id == new_task.assigned_to).first()
    return {
        "id": new_task.id, "title": new_task.title, "status": new_task.status,
        "assigned_to": new_task.assigned_to,
        "assigned_username": assigned_user.username if assigned_user else "Unknown",
        "project_id": new_task.project_id
    }

@app.get("/tasks")
def get_tasks(db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    now = datetime.utcnow()
    tasks = db.query(models.Task).all() if user["role"] == "admin" \
        else db.query(models.Task).filter(models.Task.assigned_to == user["user_id"]).all()
    result = []
    for t in tasks:
        u = db.query(models.User).filter(models.User.id == t.assigned_to).first()
        p = db.query(models.Project).filter(models.Project.id == t.project_id).first()
        result.append({
            "id": t.id, "title": t.title, "status": t.status,
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "assigned_to": t.assigned_to,
            "assigned_username": u.username if u else "Unassigned",
            "project_id": t.project_id,
            "project_name": p.name if p else "Unknown",
            "overdue": bool(t.due_date and t.due_date < now and t.status != "completed")
        })
    return result

@app.put("/tasks/{task_id}")
def update_task_status(task_id: int, update: schemas.TaskStatusUpdate = None, db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if user["role"] != "admin" and task.assigned_to != user["user_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    new_status = update.status if update and update.status in ("pending", "completed") else "completed"
    task.status = new_status
    db.commit()
    msg = "Task completed" if new_status == "completed" else "Task marked pending"
    return {"msg": msg, "id": task.id, "status": task.status}

@app.delete("/tasks/{task_id}")
def delete_task(task_id: int, db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete tasks")
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task)
    db.commit()
    return {"msg": "Task deleted"}

# ---------- DASHBOARD ----------
@app.get("/dashboard")
def dashboard(db: Session = Depends(get_db), token: str = Header(...)):
    user = get_current_user(token)
    now = datetime.utcnow()
    tasks = db.query(models.Task).all() if user["role"] == "admin" \
        else db.query(models.Task).filter(models.Task.assigned_to == user["user_id"]).all()
    result = {
        "total": len(tasks),
        "completed": len([t for t in tasks if t.status == "completed"]),
        "pending": len([t for t in tasks if t.status != "completed"]),
        "overdue": len([t for t in tasks if t.due_date and t.due_date < now and t.status != "completed"])
    }
    if user["role"] != "admin":
        recent_tasks = []
        for t in sorted(tasks, key=lambda x: (x.due_date is None, x.due_date))[:5]:
            p = db.query(models.Project).filter(models.Project.id == t.project_id).first()
            recent_tasks.append({
                "id": t.id, "title": t.title, "status": t.status,
                "project_name": p.name if p else "Unknown",
                "due_date": t.due_date.isoformat() if t.due_date else None,
                "overdue": bool(t.due_date and t.due_date < now and t.status != "completed")
            })
        result["recent_tasks"] = recent_tasks
    return result

# ---------- CHANGE PASSWORD ----------
@app.put("/change-password")
def change_password(data: schemas.ChangePassword, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == data.username).first()
    if not user or not auth.verify_password(data.current_password, user.password):
        raise HTTPException(status_code=401, detail="Invalid username or current password")
    user.password = auth.hash_password(data.new_password)
    db.commit()
    return {"msg": "Password updated successfully"}

# ---------- CSV EXPORT (admin only) ----------
@app.get("/export/csv")
def export_csv(db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    admin_only(current)
    output = io.StringIO()
    writer = csv.writer(output)
    now = datetime.utcnow()
    projects = db.query(models.Project).all()
    for idx, project in enumerate(projects):
        tasks = db.query(models.Task).filter(models.Task.project_id == project.id).all()
        owner = db.query(models.User).filter(models.User.id == project.owner_id).first()
        total = len(tasks)
        completed_c = len([t for t in tasks if t.status == "completed"])
        pending_c = total - completed_c
        overdue_c = len([t for t in tasks if t.due_date and t.due_date < now and t.status != "completed"])
        writer.writerow([f"PROJECT: {project.name}", f"Owner: {owner.username if owner else 'Unknown'}",
                         f"Total Tasks: {total}", f"Completed: {completed_c}",
                         f"Pending: {pending_c}", f"Overdue: {overdue_c}"])
        writer.writerow(["Task ID", "Task Title", "Assigned To", "Status",
                         "Due Date", "Overdue", "Completed"])
        for t in tasks:
            u = db.query(models.User).filter(models.User.id == t.assigned_to).first()
            is_overdue = bool(t.due_date and t.due_date < now and t.status != "completed")
            writer.writerow([
                t.id, t.title, u.username if u else "Unassigned", t.status,
                t.due_date.strftime("%Y-%m-%d %H:%M") if t.due_date else "No due date",
                "YES" if is_overdue else "No",
                "YES" if t.status == "completed" else "No"
            ])
        if idx < len(projects) - 1:
            writer.writerow([])
            writer.writerow([])
    output.seek(0)
    filename = f"taskpro_export_{now.strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ---------- XLSX EXPORT (admin only, multi-sheet) ----------
@app.get("/export/xlsx")
def export_xlsx(db: Session = Depends(get_db), token: str = Header(...)):
    current = get_current_user(token)
    admin_only(current)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    wb = openpyxl.Workbook()
    now = datetime.utcnow()
    projects = db.query(models.Project).all()
    all_users = {u.id: u for u in db.query(models.User).all()}

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="3C3C8A")
    subheader_fill = PatternFill("solid", fgColor="6C63FF")
    alt_fill = PatternFill("solid", fgColor="F5F5FF")
    overdue_fill = PatternFill("solid", fgColor="FFEAEA")
    done_fill = PatternFill("solid", fgColor="EAFFEA")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD")
    )

    def style_header_row(ws, row_num, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.row_dimensions[1].height = 30
    ws_sum.merge_cells("A1:G1")
    tc = ws_sum["A1"]
    tc.value = f"TaskPro Export — {now.strftime('%d %b %Y %H:%M')} UTC"
    tc.font = Font(bold=True, size=14, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="3C3C8A")
    tc.alignment = center
    ws_sum.append([])
    sum_headers = ["Project", "Owner", "Total Tasks", "Completed", "Pending", "Overdue", "Progress %"]
    ws_sum.append(sum_headers)
    style_header_row(ws_sum, 3, len(sum_headers))
    for pidx, project in enumerate(projects):
        tasks = db.query(models.Task).filter(models.Task.project_id == project.id).all()
        owner = all_users.get(project.owner_id)
        total = len(tasks); completed = len([t for t in tasks if t.status == "completed"])
        pending = total - completed
        overdue = len([t for t in tasks if t.due_date and t.due_date < now and t.status != "completed"])
        pct = round((completed / total * 100) if total > 0 else 0)
        row = [project.name, owner.username if owner else "Unknown", total, completed, pending, overdue, f"{pct}%"]
        ws_sum.append(row)
        r = ws_sum.max_row
        fill = alt_fill if pidx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(row) + 1):
            cell = ws_sum.cell(row=r, column=col)
            cell.fill = fill; cell.alignment = left; cell.border = thin
    auto_width(ws_sum)

    for project in projects:
        tasks = db.query(models.Task).filter(models.Task.project_id == project.id).all()
        owner = all_users.get(project.owner_id)
        sheet_name = project.name[:28].replace("/","-").replace("\\","-").replace("?","").replace("*","").replace("[","").replace("]","")
        ws = wb.create_sheet(title=sheet_name)
        ws.row_dimensions[1].height = 28
        ws.merge_cells("A1:H1")
        pc = ws["A1"]
        pc.value = f"Project: {project.name}"
        pc.font = Font(bold=True, size=13, color="FFFFFF")
        pc.fill = subheader_fill; pc.alignment = center
        total = len(tasks); completed_count = len([t for t in tasks if t.status == "completed"])
        overdue_count = len([t for t in tasks if t.due_date and t.due_date < now and t.status != "completed"])
        ws.append([f"Owner: {owner.username if owner else 'Unknown'}", f"Total: {total}",
                   f"Completed: {completed_count}", f"Pending: {total - completed_count}",
                   f"Overdue: {overdue_count}", f"Progress: {round((completed_count/total*100) if total else 0)}%"])
        meta_row = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=meta_row, column=col)
            c.font = Font(bold=True, size=10, color="3C3C8A")
            c.fill = PatternFill("solid", fgColor="E8E8FF"); c.alignment = left
        ws.append([])
        task_headers = ["Task ID", "Task Title", "Assigned To", "User Role", "Status", "Due Date", "Overdue?", "Completed?"]
        ws.append(task_headers)
        style_header_row(ws, ws.max_row, len(task_headers))
        for tidx, t in enumerate(tasks):
            u = all_users.get(t.assigned_to)
            is_overdue = bool(t.due_date and t.due_date < now and t.status != "completed")
            is_done = t.status == "completed"
            row = [t.id, t.title, u.username if u else "Unassigned", u.role if u else "-",
                   t.status.capitalize(),
                   t.due_date.strftime("%Y-%m-%d %H:%M") if t.due_date else "No due date",
                   "YES" if is_overdue else "No", "YES" if is_done else "No"]
            ws.append(row)
            r = ws.max_row
            row_fill = done_fill if is_done else (overdue_fill if is_overdue else (alt_fill if tidx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")))
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = row_fill; cell.alignment = left; cell.border = thin
        auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    filename = f"taskpro_export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
